import traceback

from common.http.Format import Format
from common.excel.Array import Array
from openpyxl.styles import PatternFill
from time import sleep
from jsonschema import validate
import re, chardet, os, json, datetime, time, demjson3, xmltodict

'''
@有些看起来没有用到的库是为了表达式准备的,所以不要删，其他类也一样
@author: dujianxiao
'''


class Write(Format, Array):

    def write(self, model, row, sheet, bookRes, sheetRes, fileRes, currentItera, Iteration):
        """
        写入接口请求结果
        :param model:模式(普通,简洁)
        :param row:行号
        :param sheet:用例文件
        :param bookRes:用例结果文件
        :param sheetRes:用例结果文件
        :param fileRes:用例结果文件
        :param currentItera:第n次迭代，从0计数
        :param Iteration:迭代次数
        """
        resp = []
        skipDict = []
        resultDict = []
        status = '成功'
        dict = {}
        DBExc = []
        iteraValue = self.getValue(fileRes, sheet, row, self.IterationCol)
        url = str(self.getValue(fileRes, sheet, row, self.urlCol))
        url = self.repRel(url)
        url = self.repVar(str(url))
        className = str(self.getValue(fileRes, sheet, row, self.nameCol))
        className = self.repRel(className)
        className = self.repVar(str(className))
        if isinstance(iteraValue, int) == False and iteraValue != '':
            skipDict = self.setSkip(sheet, row, bookRes, sheetRes, fileRes, '迭代异常', currentItera, Iteration, '')
            status = '异常'
            duration = '--'
            resultDict = []
            DBExc = []
        else:
            # 数据库的连接和关闭的时机有严格的逻辑，勿动。
            conn = self.getConn(fileRes, sheet, row)
            r, duration, msg = self.checkFormat(fileRes, sheet, row, conn)
            print(r, duration, msg)
            try:
                resp.append(str(r.text))
                # 普通模式下打印接口响应
                # 如果接口返回中含有html元素,使用append方式显示的是渲染后的结果,使用insertPlainText显示的是html原文
                # insertPlainText性能很差,很容易导致页面卡死进而导致程序崩溃
                if model == '普  通':
                    form, ss = self.getType(r)
                    num = len(ss) // 1000 + 1
                    for i in range(num):
                        if form == 'xml':
                            self.consoleFunc('')
                            self.console.insertPlainText(ss[i * 1000:(i + 1) * 1000])
                        elif form == 'json' or form == 'jsonp':
                            self.consoleFunc('black', ss[i * 1000:(i + 1) * 1000])
                        # 由于qt性能的原因，每1000个字符暂停100毫秒，100毫秒是一个经验值
                        time.sleep(0.1)
                    # 解决普通模式下客户文字错位和文字颜色与预期不问题
                    time.sleep(0.1)
            except Exception as e:
                print(e)
            if '异常' in str(msg):
                if '数据库异常' not in str(msg):
                    url = self.rep(fileRes, sheet, row, conn, url)
                    className = self.rep(fileRes, sheet, row, conn, className)
                skipDict = self.setSkip(sheet, row, bookRes, sheetRes, fileRes, msg, currentItera, Iteration, conn)
                status = '异常'
            else:
                url = self.rep(fileRes, sheet, row, conn, url)
                # 取校验数据、预期结果的原始值和结果值
                checkRes = self.checkRes(r, fileRes, sheet, row, conn)
                initRes = self.expResultInit(fileRes, sheet, row, conn)
                check = self.check(fileRes, sheet, row, conn)
                result = self.expResult(fileRes, sheet, row, conn)
                # 在数据恢复之前进行三者替换
                statusCode = self.getArray(fileRes, sheet, row, self.statusCodeCol, self.expressionCol)
                resHeader = self.getArray(fileRes, sheet, row, self.resHeaderCol, self.statusCodeCol)
                res = self.getArray(fileRes, sheet, row, self.resTextCol, self.resHeaderCol)
                expression = self.getArray(fileRes, sheet, row, self.expressionCol, self.statusCol)

                statusCode = [self.repAll(str(item), fileRes, sheet, row, conn) for item in statusCode]
                resHeader = [self.repAll(str(item), fileRes, sheet, row, conn) for item in resHeader]
                res = [self.repAll(str(item), fileRes, sheet, row, conn) for item in res]
                expression = [self.repAll(str(item), fileRes, sheet, row, conn) for item in expression]
                # 数据恢复之前把所有数据库相关的操作处理完
                resMsg = self.restore(fileRes, sheet, row, conn)
                # 数据库恢复部分的SQL异常
                if resMsg:
                    skipDict = self.setSkip(sheet, row, bookRes, sheetRes, fileRes, resMsg, currentItera, Iteration,
                                            conn)
                    status = '异常'
                else:
                    # 完全没有异常了再执行setResult
                    resultDict = self.setResult(row, bookRes, sheetRes, fileRes, checkRes, check, result,
                                                initRes, r, duration, res, resHeader, statusCode, expression,
                                                currentItera, Iteration)
                    if len(resultDict) > 0:
                        status = '失败'
            try:
                if conn == [[]]:
                    pass
                else:
                    if '数据库异常' not in str(conn):
                        conn[0][0].close()
            except Exception as e:
                print(e)
        # 只统计最后一次的结果
        if currentItera == Iteration - 1:
            # 信息存入字典，用于html测试报告
            dict['className'] = className
            dict['url'] = url
            dict['method'] = self.getValue(fileRes, sheet, row, self.methodCol)
            dict['param'] = self.getValue(fileRes, sheet, row, self.paramCol)
            dict['header'] = self.getValue(fileRes, sheet, row, self.headerCol)
            dict['duration'] = duration
            dict['resp'] = resp
            dict['status'] = status
            dict['log'] = skipDict + resultDict + DBExc
            return dict
        else:
            return []

    def getType(self, r):
        """
        获取接口响应类型：xml,json,jsonp
        :param r:接口响应对象
        """
        form = ''
        ss = ''
        try:
            try:
                # 某些接口返回值是html格式,会出现大量的转义字符,使用loads进行反序列化
                ss = str(json.loads(r.text))
                form = 'json'
            except Exception as e:
                print(e)
                try:
                    # 又由于有些接口返回值不是json格式,不能loads,所以如果反序列化失败即不再进行反序列化
                    eval('json.dumps(xmltodict.parse(r.text))')
                    ss = str(r.text)
                    form = 'xml'
                except Exception as e:
                    print(e)
                    try:
                        # 格式为jsonp
                        eval('json.loads(re.search("^[^(]*?\((.*)\)[^)]*$",r.text,re.S).group(1))')
                        #                         ss=str(json.loads(re.match(".*?({.*}).*",r.text,re.S).group(1)))
                        ss = str(r.text)
                        form = 'jsonp'
                    except Exception as e:
                        print(e)
        except:
            if str(r) != '':
                self.consoleFunc('red', str(r))
        return form, ss

    def analyFunc(self, fileRes, row, sheetName, sheet):
        """
        解析JSON
        :param fileRes:用例结果文件
        :param row:行号
        :param sheetName:页签名
        :param sheet:用例文件
        """

        # JSON解析中不对异常情况进行处理，如有异常直接解析失败
        className = str(self.getValue(fileRes, sheet, row - 1, self.nameCol))
        self.setFlag(sheetName, row + 1, className, '解析开始')
        self.consoleFunc('green', str(row + 1) + ' ' + str(self.getValue(fileRes, sheet, row, self.nameCol)))
        try:
            conn = self.getConn(fileRes, sheet, row)
            s1, s2 = self.jsonFormat(fileRes, sheet, row, conn)
            if s2 == '解析失败':
                self.consoleFunc('red', '解析失败.')
            else:
                for i in range(len(s1)):
                    self.consoleFunc('black', f"{s1[i]}:{s2[i]}")
                    time.sleep(0.001)
            self.consoleFunc('black')
        except Exception as e:
            print(e)
            self.consoleFunc('red', '解析失败.')
        self.setFlag(sheetName, row + 1, className, '解析结束')

    def setSkip(self, sheet, row, bookRes, sheetRes, fileRes, msg, currentItera, Iteration, conn):
        """
        如果数据合法性校验不通过则调用此方法
        :param sheet:用例文件
        :param row:行号
        :param bookRes:用例结果文件
        :param sheetRes:用例结果文件
        :param fileRes:用例结果文件
        :param msg: 接口返回的异常信息
        :param Iteration:迭代次数
        :param currentItera:第n次迭代，从0计数
        :param conn:数据库连接对象
        """
        skipDict = []
        blue = self.setCellStyle(7)
        if '迭代异常' in str(msg):
            self.consoleFunc('red', '迭代次数只能为空或非负整数')
            self.status3 = self.status3 + 1
            iteraValue = self.getValue(fileRes, sheet, row, self.IterationCol)
            skipDict.append("迭代次数异常:" + str(iteraValue))
            # 标识结果为：skip，并设背景为蓝色
            if fileRes.endswith('xls'):
                sheetRes.write(row, self.IterationCol, iteraValue, blue)
                sheetRes.write(row, self.statusCol, 'skip', blue)
            elif fileRes.endswith('xlsx'):
                self.setValueColor(sheetRes, row + 1, self.IterationCol, iteraValue, "blue")
                self.setValueColor(sheetRes, row + 1, self.statusCol, 'skip', "blue")
        else:
            if '数据库异常' in str(msg):
                if msg[0][1] == self.DBCol and msg[1] == []:
                    # 有sql未选择数据库
                    self.getError(str('有sql语句而没有连接数据库'))
                    self.consoleFunc('red', '有sql语句而没有连接数据库')
                    self.consoleFunc('red', str(msg[0]))
                elif msg[0][1] == self.DBCol and msg[1] != []:
                    # 数据库连接类的异常
                    err = msg[1][0]
                    self.consoleFunc('red', err)
                    self.consoleFunc('red', str(msg[0]))
                else:
                    # sql执行异常
                    for i in range(1, len(msg[0])):
                        # exceValue = self.getValue(fileRes, sheet, row, int(msg[0][i]))
                        self.consoleFunc('red', msg[1][i - 1])
                    self.consoleFunc('red', str(msg[0]))
                # 异常信息存skipDict用于html测试报告
                skipDict.append(str(msg[0]))
                if msg[1]:
                    skipDict.append(str(msg[1]))
            else:
                skipDict.append(str(msg))
                self.getToLog(str(msg))
                for i in range(1, len(msg)):
                    exceValue = self.getValue(fileRes, sheet, row, int(msg[i]))
                    exceValue = self.repAll(str(exceValue), fileRes, sheet, row, conn)
                    self.consoleFunc('red', exceValue)
                    skipDict.append(exceValue)
                    self.getToLog(exceValue)
                self.consoleFunc('red', str(msg))
            if currentItera == Iteration - 1:
                self.status3 = self.status3 + 1
            # 标识结果为：skip，并设背景为蓝色
            if fileRes.endswith('xls'):
                sheetRes.write(row, self.statusCol, 'skip', blue)
            elif fileRes.endswith('xlsx'):
                self.setValueColor(sheetRes, row + 1, self.statusCol, 'skip', "blue")
            # 去掉异常信息的数组
            if '数据库异常' in str(msg):
                newArr = self.filterArr(msg[0], '异常')
            else:
                newArr = self.filterArr(msg, '异常')
            # 标识合法性校验不通过的单元格为蓝色
            for item in newArr:
                if fileRes.endswith('xls'):
                    sheetRes.write(row, int(item), self.getValue(fileRes, sheet, row, int(item)), blue)
                elif fileRes.endswith('xlsx'):
                    self.setValueColor(sheetRes, row + 1, int(item), self.getValue(fileRes, sheet, row, int(item)),
                                       "blue")
        bookRes.save(fileRes)
        return skipDict

    def setResult(self, row, bookRes, sheetRes, fileRes, checkRes, check, result, initRes, r, duration, res,
                  resHeader, statusCode, expression, currentItera, Iteration):
        """
        数据合法性校验通过后调用此方法，校验各字段的值是否正确
        :param Iteration:
        :param currentItera:
        :param expression:
        :param statusCode:
        :param res:
        :param resHeader:
        :param row:行号
        :param bookRes:用例结果文件
        :param sheetRes:用例结果文件
        :param fileRes:用例结果文件
        :param checkRes:校验字段结果数组+文件数组
        :param check: 校验字段数组－－原值
        :param result: 预期结果值数组
        :param initRes: 预期结果数组－－原值
        :param r:接口响应对象
        :param duration:接口响应时间
        """
        resultDict = []
        red = self.setCellStyle(2)
        green = self.setCellStyle(3)
        status = 0
        # 预置结果为true，后面如果有错误再修改结果
        if fileRes.endswith('xls'):
            sheetRes.write(row, self.timeCol, duration)
            sheetRes.write(row, self.statusCol, 'true', green)
        elif fileRes.endswith('xlsx'):
            self.setValueColor(sheetRes, row + 1, self.timeCol, duration, "")
            self.setValueColor(sheetRes, row + 1, self.statusCol, 'true', "green")
        # 校验预期结果，精确匹配
        for j in range(len(check)):
            if str(checkRes[j]) != str(result[j]):  #
                if fileRes.endswith('xls'):
                    sheetRes.write(row, self.part101Col + j,
                                   f"{check[j]}-->{checkRes[j]}:{result[j]}", red)
                    sheetRes.write(row, self.section101Col + j, str(initRes[j]), red)
                    sheetRes.write(row, self.statusCol, 'false', red)
                elif fileRes.endswith('xlsx'):
                    self.setValueColor(sheetRes, row + 1, self.part101Col + j,
                                       f"{check[j]}-->{checkRes[j]}:{result[j]}", "red")
                    self.setValueColor(sheetRes, row + 1, self.section101Col + j, str(initRes[j]), "red")
                    self.setValueColor(sheetRes, row + 1, self.statusCol, 'false', "red")
                self.consoleFunc('red', f"{check[j]}:实际结果:{checkRes[j]}-->预期结果:{result[j]}")
                resultDict.append(f"{check[j]}:实际结果:{checkRes[j]}-->预期结果:{result[j]}")
                status = 1
        # 响应断言
        for i in range(len(res)):
            if res[i] in r.text:
                pass
            else:
                if fileRes.endswith('xls'):
                    sheetRes.write(row, self.statusCol, 'false', red)
                    sheetRes.write(row, self.resTextCol + i, res[i], red)
                elif fileRes.endswith('xlsx'):
                    self.setValueColor(sheetRes, row + 1, self.statusCol, 'false', "red")
                    self.setValueColor(sheetRes, row + 1, self.resTextCol + i, res[i], "red")
                self.consoleFunc('red', f"响应断言失败:{res[i]}")
                resultDict.append(f"响应断言失败:{res[i]}")
                status = 1
        # 校验响应头，模糊匹配
        for i in range(len(resHeader)):
            if str(resHeader[i]) == '':
                pass
            elif str(resHeader[i]) not in str(r.headers):
                if fileRes.endswith('xls'):
                    sheetRes.write(row, self.resHeaderCol + i, str(resHeader[i]), red)
                    sheetRes.write(row, self.statusCol, 'false', red)
                elif fileRes.endswith('xlsx'):
                    self.setValueColor(sheetRes, row + 1, self.resHeaderCol + i, str(resHeader[i]), "red")
                    self.setValueColor(sheetRes, row + 1, self.statusCol, 'false', "red")
                self.consoleFunc('red', f"响应头断言失败:{resHeader[i]}")
                resultDict.append(f"响应头断言失败:{resHeader[i]}")
                status = 1
        # 校验响应码，精确匹配
        for i in range(len(statusCode)):
            if str(statusCode[i]) == '':
                pass
            elif str(statusCode[i]) != str(r.status_code):
                if fileRes.endswith('xls'):
                    sheetRes.write(row, self.statusCodeCol + i,
                                   f"{statusCode[i]}-->{r.status_code}:{statusCode[i]}", red)
                    sheetRes.write(row, self.statusCol, 'false', red)
                elif fileRes.endswith('xlsx'):
                    self.setValueColor(sheetRes, row + 1, self.statusCodeCol + i,
                                       f"{statusCode[i]}-->{r.status_code}:{statusCode[i]}",
                                       "red")
                    self.setValueColor(sheetRes, row + 1, self.statusCol, 'false', "red")
                self.consoleFunc('red', f"响应码断言失败:实际结果:{r.status_code}-->预期结果:{statusCode[i]}")
                resultDict.append(f"响应码断言失败:实际结果:{r.status_code}-->预期结果:{statusCode[i]}")
                status = 1
        # 校验表达式
        js = self.getResType(r)
        for i in range(len(expression)):
            expreFlag = True
            expression[i] = str(expression[i]).replace("r.json()", js)
            if expression[i] == '':
                expreFlag = True
            else:
                expreFlag = eval(expression[i])
            if expreFlag:
                pass
            else:
                if fileRes.endswith('xls'):
                    sheetRes.write(row, self.expressionCol + i, str(expression[i]), red)
                    sheetRes.write(row, self.statusCol, 'false', red)
                elif fileRes.endswith('xlsx'):
                    self.setValueColor(sheetRes, row + 1, self.expressionCol + i, str(expression[i]), "red")
                    self.setValueColor(sheetRes, row + 1, self.statusCol, 'false', "red")
                self.consoleFunc('red', f"表达式断言失败:{expression[i]}")
                resultDict.append(f"表达式断言失败:{expression[i]}")
                status = 1
        if currentItera == Iteration - 1:
            if status == 1:
                self.status2 = self.status2 + 1
            else:
                self.status1 = self.status1 + 1
        bookRes.save(fileRes)
        return resultDict

    def setValueColor(self, sheetRes, row, column, value, color):
        """
        写入值并设置背景色
        :param sheetRes:
        :param row:行号
        :param column:列号
        :param value:写入单元格的值
        :param color:单元格背景色
        """
        sheetRes.cell(row=row, column=column, value=value)
        color_fill = PatternFill("solid", fgColor=color)
        sheetRes.cell(row, column).fill = color_fill

    def run(self, model, n, sheetName, sheet, nrows, bookRes, sheetRes, fileRes, allRows):
        """
        执行－－单行执行或全量执行（无参数）
        :param model:模式(普通,简洁)
        :param n:行号
        :param sheetName:页签名
        :param sheet:用例文件
        :param nrows:行数
        :param bookRes:用例结果文件
        :param sheetRes: 用例结果文件
        :param fileRes:用例结果文件
        :param allRows:全部用例数
        """
        testResult = []
        dict = {}
        if n == '':
            # 全量执行
            self.consoleFunc('blue', f"【{sheetName}】", 'size=4')
            self.consoleFunc('black')
            for row in range(3, nrows + 1):
                className = str(self.getValue(fileRes, sheet, row - 1, self.nameCol))
                Iteration = self.getValue(fileRes, sheet, int(row) - 1, self.IterationCol)
                if isinstance(Iteration, int):
                    for i in range(Iteration):
                        print(row)
                        self.consoleFunc('green', str(row) + ' ' + className)
                        self.setFlag(sheetName, row, className, '请求开始')
                        if i == Iteration - 1:
                            testResult.append(
                                self.write(model, row - 1, sheet, bookRes, sheetRes, fileRes, i, Iteration))
                        else:
                            self.write(model, row - 1, sheet, bookRes, sheetRes, fileRes, i, Iteration)
                        self.successNum.setText(str(self.status1))
                        self.failNum.setText(str(self.status2))
                        self.skipNum.setText(str(self.status3))
                        self.result.setText(f"{self.status1 + self.status2 + self.status3}/{allRows}")
                        self.setFlag(sheetName, row, className, '请求结束')
                else:
                    print(row)
                    self.consoleFunc('green', str(row) + ' ' + className)
                    self.setFlag(sheetName, row, className, '请求开始')
                    testResult.append(self.write(model, row - 1, sheet, bookRes, sheetRes, fileRes, 0, 1))
                    self.successNum.setText(str(self.status1))
                    self.failNum.setText(str(self.status2))
                    self.skipNum.setText(str(self.status3))
                    self.result.setText(f"{self.status1 + self.status2 + self.status3}/{allRows}")
                    self.setFlag(sheetName, row, className, '请求结束')
                self.consoleFunc('black')
        else:
            # 单个或多个
            className = str(self.getValue(fileRes, sheet, n - 1, self.nameCol))
            Iteration = self.getValue(fileRes, sheet, int(n) - 1, self.IterationCol)
            if isinstance(Iteration, int):
                for i in range(Iteration):
                    print(n)
                    self.consoleFunc('green', str(n) + ' ' + className)
                    self.setFlag(sheetName, n, className, '请求开始')
                    if i == Iteration - 1:
                        testResult.append(self.write(model, n - 1, sheet, bookRes, sheetRes, fileRes, i, Iteration))
                    else:
                        self.write(model, n - 1, sheet, bookRes, sheetRes, fileRes, i, Iteration)
                    self.successNum.setText(str(self.status1))
                    self.failNum.setText(str(self.status2))
                    self.skipNum.setText(str(self.status3))
                    self.result.setText(f"{self.status1 + self.status2 + self.status3}/{allRows}")
                    self.setFlag(sheetName, n, className, '请求结束')
            else:
                print(n)
                self.consoleFunc('green', str(n) + ' ' + className)
                self.setFlag(sheetName, n, className, '请求开始')
                testResult.append(self.write(model, n - 1, sheet, bookRes, sheetRes, fileRes, 0, 1))
                self.successNum.setText(str(self.status1))
                self.failNum.setText(str(self.status2))
                self.skipNum.setText(str(self.status3))
                self.result.setText(f"{self.status1 + self.status2 + self.status3}/{allRows}")
                self.setFlag(sheetName, n, className, '请求结束')
            self.consoleFunc('black')
        # 用于html测试报告
        dict['testAll'] = self.status1 + self.status2 + self.status3
        dict['testPass'] = self.status1
        dict['testFail'] = self.status2
        dict['testSkip'] = self.status3
        return dict, testResult

    def setFlag(self, sheetName, row, className, content):
        self.getToLog(f"{'☆' * 20}【{sheetName}】第{row + 1}个接口【{className}】{content}{'☆' * 20}")
