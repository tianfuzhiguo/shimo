from xlwt import Pattern
from xlwt import XFStyle 
from common.utils.Log import getError
from common.utils.Log import getToLog
import xlrd
import openpyxl
import cx_Oracle
import pymysql
import pymssql

'''
@author: dujianxiao
'''


'''
@连接数据库 
@param file:用例文件
@param sheet:  
@param row:行号 
@param column:列号
@param userParams:用户变量
@param userParamsValue:用户变量值
'''
def getConn(file,sheet,row,column,userParams,userParamsValue):
    DB=getValue(file,sheet,row,column[22])
    DB=repVar(str(DB),userParams,userParamsValue)
    if DB=='':
        return [[]]
    else:
        try: 
            #DB1,DB2,DB3是配置文件中预置的3个数据库
            if DB=="DB1":
                DB="${DB1}"
            if DB=="DB2":
                DB="${DB2}"
            if DB=="DB3":
                DB="${DB3}"
            DB=repVar(str(DB),userParams,userParamsValue)
            conn=eval(DB)
        except Exception as e:
            print(e)
            getError(str(e))
            return [['数据库异常',column[22]],[str(e)]]
        return [[conn]] 
    
'''
@deprecated: 去除Arr中含有word的字符串
@param arr:数组
@param word:关键字  
'''
def filterArr(arr, word):
    return [item for item in arr if str(word) not in str(item)]

'''
@deprecated: 读取用例文件,支持.xls和.xlsx格式
@param file:用例文件 
'''              
def readExcel(file):
    data=''
    if file[-4:]=='.xls':
        data = xlrd.open_workbook(file)
    elif file[-5:]=='.xlsx':
        data = openpyxl.load_workbook(file)
    return data

'''
@deprecated: 获取用例文件中的全部页签名称
@param file:用例文件
@param data:  
'''
def getSheetNames(file,data):
    sheetNames=''
    if file[-4:]=='.xls':
        sheetNames = data.sheet_names()
    elif file[-5:]=='.xlsx':
        sheetNames = data.get_sheet_names()
    return sheetNames

'''
@deprecated: 用于查找特定字符串所在的列号
@param file:用例文件
@param sheet:  
@param ncols:列数 
@param field:关键字 
'''
def findStr(file,sheet,ncols,field):
    try:
        if file[-4:]=='.xls':
            for i in range(1, ncols):
                re = sheet.cell(1, i).value
                if(re == field):
                    return i
        elif file[-5:]=='.xlsx':
            for i in range(1, ncols):
                re = sheet.cell(row=2, column=i).value
                if(re == field):
                    return i
    except Exception as e:
        print(e)
        return '未查找到字符串：' + field
    
'''
@deprecated: 获取某行某列的值
@param file:用例文件
@param sheet:  
@param row:行号
@param column:列号
'''
def getValue(file,sheet,row,column):
    if file[-4:]=='.xls':
        ctype = sheet.cell(row, column).ctype  # 表格的数据类型
        flag = sheet.cell_value(row, column)
        if ctype == 2 and flag % 1 == 0:  # 如果是整形
            flag = int(flag)
        else:
            flag = flag
    elif file[-5:]=='.xlsx':
        flag = sheet.cell(row=row+1, column=column).value
        if flag==None:
            flag=''
    return flag

'''       
@deprecated: 获取某行start到end之间的数组－－原值
@param file:用例文件
@param sheet:  
@param row:行号
@param start:
@param end:   
'''
def getArray(file,sheet,row,start,end):
    result = []
    if file[-4:]=='.xls':
        for column in range(start, end):
            ctype = sheet.cell(row, column).ctype  # 表格的数据类型
            cell = sheet.cell_value(row, column)
            if ctype == 2 and cell % 1 == 0:  # 如果是整形
                cell = int(cell)
            result.append(str(cell))  
    elif file[-5:]=='.xlsx':
        for column in range(start, end):
            value = sheet.cell(row=row+1, column=column).value
            if value==None:
                value=''
            result.append(str(value)) 
    return result

'''
@deprecated: 取异常数据中的真实值
@param file:用例文件
@param sheet:
@param row:行号   
@param field:关键字
@param msg:异常数组
'''
def getInitArray(file,sheet,row,field,msg):
    return [getValue(file,sheet,row,int(item)) for item in msg[1:]] if field in str(msg) else ''
'''
@deprecated: 获取sql结果数组
@param file:用例文件
@param sheet:  
@param row:行号
@param conn:数据库连接对象
@param start: 
@param end:  
'''
def getSqlResultArray(file,sheet,row, conn, start, end):
    data = []
    sqlArray = getArray(file,sheet,row, start, end)
    column=start
    '''
    @在此之前已经校验过sql不全为空而数据库为空的情况
    @所以如果此时数据库为空，说明sql全为空
    '''
    if conn==[[]]:
        return sqlArray
    else:
        cursor = conn[0][0].cursor()
        for item in sqlArray:
            if(item == ''):
                data.append('')
            else:
                cursor.execute(item)
                dd = cursor.fetchone()
                data.append(None if dd==None else dd[0])
            column=column+1
        cursor.close()
        conn[0][0].commit()
        return data

'''
@有SQL则数据库不允许为空
'''
def DBExists(file,sheet,row,column,conn):
    allSql = getArray(file,sheet,row,column[7],column[8]) + getArray(file,sheet,row,column[9],column[10]) + getArray(file,sheet,row,column[16],column[19])
    '''
    @如果sql不全为空而数据库连接为空，则返回数据库异常，否则返回空数组
    '''
    return [['数据库异常', column[22]],[]] if ''.join(allSql)!='' and conn==[[]] else []

'''
@deprecated: 此方法仅用作验证SQL句是否正确
@param file:用例文件
@param sheet:  
@param row:行号 
@param column:列号
@param userParams:用户变量
@param userParamsValue:用户变量值
@param userVar:接口变量数组 
@param userVarValue:接口变量值数组 
'''
def sqlExcept(file,sheet,row,column,userParams,userParamsValue,userVar,userVarValue,conn):
    conn=getConn(file,sheet,row,column,userParams,userParamsValue)
    msg = ['数据库异常']
    SqlMsg = []
    sqlArray = getArray(file,sheet,row,column[7],column[8]) + getArray(file,sheet,row,column[9],column[10]) + getArray(file,sheet,row,column[18],column[19])
    
    for item in sqlArray:
        item=repVar(str(item),userParams,userParamsValue)
        item=repRel(row,userVar,userVarValue,str(item))
        
    if conn==[[]]:
        return []
    else:   
        column1=column[7] 
        cursor = conn[0][0].cursor()
        for item in sqlArray:
            if item != '':
                try:
                    '''                   
                    #这三部分的SQL只能是查询语句
                    '''
                    if (str(item).lower()).replace(' ', '').startswith('select')!=True:
                        msg.append(column1)
                        SqlMsg.append(column1)
                    else:
                        cursor.execute(item)   
                except Exception as e:
                    msg.append(column1)
                    getError(str(e))
                    SqlMsg.append(str(e))
#                     msg.append(str(e))
            if column1==column[8]-1:
                column1=column[9]-1
            if column1==column[10]-1:
                column1=column[18]-1
            column1=column1+1
        cursor.close()
        conn[0][0].close()

        return [msg,SqlMsg] if len(msg)>1 else []
    
'''    
@deprecated: 数据初始化
@param file:用例文件
@param sheet:  
@param row:行号
@param conn:数据库连接对象 
@param param: 数据库连接对象 
@param userParams:用户变量
@param userParamsValue:用户变量值
'''
def init(file,sheet,row,conn,column,userParams,userParamsValue,userVar,userVarValue):
    msg = ['数据库异常']
    SqlMsg = []
    column1=column[16]
    sqlArray = getArray(file,sheet,row,column[16],column[17])
    for item in sqlArray:
        item=repVar(str(item),userParams,userParamsValue)
        item=repRel(row,userVar,userVarValue,str(item))
    if conn ==[[]]:
        return []
    else:
        cursor = conn[0][0].cursor()
        for item in sqlArray:
            if item != '':
                try:
                    item=rep(file,sheet,row,conn,item,column,userParams,userParamsValue,userVar,userVarValue)
                    cursor.execute(item)
                    conn[0][0].commit()
                    getToLog('数据初始化：'+item)
                except Exception as e:
                    print(e)
                    msg.append(str(column1))
                    getError(str(e))
                    SqlMsg.append(str(e))
            column1=column1+1
        cursor.close()
        return [msg,SqlMsg] if len(msg)>1 else []
    
'''    
@deprecated: 数据恢复
@param file:用例文件
@param sheet:  
@param row: 行号
@param conn:数据库连接对象 
@param column:列号 
@param userParams:用户变量
@param userParamsValue:用户变量值
@param userVar:接口变量数组 
@param userVarValue:接口变量值数组 
'''
def restore(file,sheet,row,conn,column,userParams,userParamsValue,userVar,userVarValue):
    msg = ['数据库异常']
    SqlMsg = []
    column1 = column[17]
    sqlArray = getArray(file,sheet,row, column[17], column[18])
    for item in sqlArray:
        item=repVar(str(item),userParams,userParamsValue)
        item=repRel(row,userVar,userVarValue,str(item))
    if conn == [[]]:
        return []
    else:
        cursor = conn[0][0].cursor()
        for item in sqlArray:
            if item != '':
                try:
                    item=rep(file,sheet,row,conn,item,column,userParams,userParamsValue,userVar,userVarValue)
                    cursor.execute(item)
                    conn[0][0].commit()
                    getToLog('数据恢复：'+item)
                except Exception as e:
                    print(e)
                    msg.append(str(column1))
                    getError(str(e))
                    SqlMsg.append(str(e))
            column1=column1+1
        cursor.close()
        return [msg,SqlMsg] if len(msg)>1 else []

'''    
@deprecated: 动态化参数
@param file:用例文件
@param sheet:  
@param row: 行号
@param conn: 数据库连接对象
@param column:列号 
@param userParams:用户变量
@param userParamsValue:用户变量值
@param userVar:接口变量数组 
@param userVarValue:接口变量值数组 
'''
def dyparam(file,sheet,row,conn,column,userParams,userParamsValue,userVar,userVarValue):
    data=[]
    sqlArray = getArray(file,sheet,row, column[18], column[19])
    for item in sqlArray:
        item=repVar(str(item),userParams,userParamsValue)
        item=repRel(row,userVar,userVarValue,str(item))
    if conn == [[]]:
        return []
    else:
        cursor = conn[0][0].cursor()
        for item in sqlArray:
            if(item == ''):
                data.append('')
            else:
                cursor.execute(item)
                dd = cursor.fetchone()
                data.append(None if dd==None else dd[0])
                conn[0][0].commit()
        cursor.close()
        return data
        

'''
@deprecated: 用户变量替换
@param param: 参数
@param userParams:用户变量数组
@param userParamsValue:用户变量值数组 
'''
def repVar(param,userParams,userParamsValue):
    for i in range(0,len(userParams)):
        if '${'+userParams[i]+'}' in param:
            param=param.replace('${'+str(userParams[i])+'}',str(userParamsValue[i]))
    return param    

        
'''
@deprecated: 接口变量替换
@param row: 行号
@param userVar:接口变量数组 
@param userVarValue:接口变量值数组 
@param param:参数
'''
def repRel(row,userVar,userVarValue,param):
    for i in range(0, len(userVar)):
        if '${'+str(userVar[i])+'}' in str(param):
            param=param.replace('${'+str(userVar[i])+'}',str(userVarValue[i]))
    return param        

'''
@deprecated: 动态参数替换
@param file:用例文件
@param sheet:  
@param row: 行号
@param conn: 数据库连接对象
@param param: 动态参数
@param column:列号 
@param userParams:用户变量
@param userParamsValue:用户变量值
@param userVar:接口变量数组 
@param userVarValue:接口变量值数组 
'''
def rep(file,sheet,row,conn,param,column,userParams,userParamsValue,userVar,userVarValue):
    data=''
    dypArr=[]
    dypar=dyparam(file,sheet,row,conn,column,userParams,userParamsValue,userVar,userVarValue)
    if dypar==None:
        return param
    else:
        for i in range(1,len(dypar)+1):
            dypArr.append('dyparam'+str(i).zfill(3))
        for i in range(0, len(dypar)):
            if '${'+str(dypArr[i])+'}' in param:
                param=param.replace('${'+str(dypArr[i])+'}',str(dypar[i]))
    return param

'''
@三者替换
@替换用户变量、动态参数、接口变量
'''
def repAll(strValue,file,sheet,row,conn,column,userParams,userParamsValue,userVar,userVarValue):
    strValue=repVar(str(strValue),userParams,userParamsValue)
    strValue=rep(file,sheet,row,conn,strValue,column,userParams,userParamsValue,userVar,userVarValue)
    strValue=repRel(row,userVar,userVarValue,strValue)
    return strValue

'''
@格式化响应时间
@param duration:接口响应时间
'''
def getDuration(duration):
    n = 0
    l = len(duration)
    for i in range(0,l):
        if duration[i].isdigit()==True and duration[i]!='0' and i<7:
            n = i
            break
        elif duration[i].isdigit()==True and duration[i]!='0' and i>7:
            n = 6
            break
    du = duration[n:-3] 
    return du

'''
@设置单元格格式
'''
def setStyle(n):
    pattern = Pattern()
    pattern.pattern = Pattern.SOLID_PATTERN 
    pattern.pattern_fore_colour = n 
    style = XFStyle()
    style.pattern = pattern 
    return style

