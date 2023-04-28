import openpyxl
import xlrd
from openpyxl.styles import PatternFill
from xlwt import Pattern
from xlwt import XFStyle

'''
Created on 2022年1月6日
@author: dujianxiao
'''


class ExcelUtil:

    @staticmethod
    def readExcel(file):
        """
        读取用例文件,支持.xls和.xlsx格式
        @param file: 用例文件
        """
        book = ''
        if file.endswith('xls'):
            book = xlrd.open_workbook(file)
        elif file.endswith('xlsx'):
            book = openpyxl.load_workbook(file)
        return book

    @staticmethod
    def getValue(file, sheet, row, columnNum):
        """
        获取某行某列的值
        @param file: 用例文件
        @param sheet: 页签
        @param row: 行号
        @param columnNum: 列号
        """
        if file.endswith('xls'):
            ctype = sheet.cell(row, columnNum).ctype  # 表格的数据类型
            flag = sheet.cell_value(row, columnNum)
            if ctype == 2 and flag % 1 == 0:  # 如果是整形
                flag = int(flag)
            else:
                flag = flag
        elif file.endswith('xlsx'):
            flag = sheet.cell(row=row + 1, column=columnNum).value
            if flag is None:
                flag = ''
        return flag

    def findStr(self, file, sheet, field):
        """
        用于查找特定字符串所在的列号
        @param file: 用例文件
        @param sheet: 页签
        @param field: 关键字
        """
        for i in range(1, self.ncols):
            if file.endswith('xls'):
                re = sheet.cell(1, i).value
            elif file.endswith('xlsx'):
                re = sheet.cell(row=2, column=i).value
            if re == field:
                return i

    @staticmethod
    def getSheetNames(file):
        """
        获取用例文件中的全部页签名称
        @param file: 用例文件
        """
        sheetNames = ''
        book = ExcelUtil.readExcel(file)
        if file.endswith('xls'):
            sheetNames = book.sheet_names()
        elif file.endswith('xlsx'):
            sheetNames = book.get_sheet_names()
        return sheetNames

    @staticmethod
    def filterList(dataList: list, word: str):
        """
        过滤数组中含有指定字符串的元素
        @param dataList: 数组
        @param word: 关键字
        """
        return [item for item in dataList if f'{word}' not in f'{item}']

    @staticmethod
    def getList(file, sheet, row, start, end):
        """
        获取某行start到end之间的数组－－原值
        @param file: 用例文件
        @param sheet: 页签
        @param row: 行号
        @param start: 索引开始
        @param end: 索引结束
        """
        result = []
        if file.endswith('xls'):
            for column in range(start, end):
                ctype = sheet.cell(row, column).ctype  # 表格的数据类型
                cell = sheet.cell_value(row, column)
                if ctype == 2 and cell % 1 == 0:  # 如果是整形
                    cell = int(cell)
                result.append(f'{cell}')
        elif file.endswith('xlsx'):
            for column in range(start, end):
                value = sheet.cell(row=row + 1, column=column).value
                if value is None:
                    value = ''
                result.append(f'{value}')
        return result

    @staticmethod
    def setCell(n):
        """
        设置单元格格式
        """
        pattern = Pattern()
        pattern.pattern = Pattern.SOLID_PATTERN
        pattern.pattern_fore_colour = n
        style = XFStyle()
        style.pattern = pattern
        return style

    @staticmethod
    def setColor(sheetRes, row, column, value, color):
        """
        写入值并设置背景色
        @param sheetRes: 结果页签
        @param row: 行号
        @param column: 列号
        @param value: 写入单元格的值
        @param color: 单元格背景色
        """
        sheetRes.cell(row=row, column=column, value=value)
        color_fill = PatternFill("solid", fgColor=color)
        sheetRes.cell(row, column).fill = color_fill
